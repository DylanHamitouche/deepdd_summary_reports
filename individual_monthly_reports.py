import pandas as pd
import os
import matplotlib.pyplot as plt
import numpy as np
import matplotlib.gridspec as gridspec
import matplotlib.dates as mdates
import datetime as dtm
from datetime import datetime, timedelta
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
from matplotlib import rcParams

# Set the font globally
rcParams['font.family'] = 'Roboto'


# DEFINE ZONES CUT-OFFS FOR REDCAP QUESTIONNAIRES
phq9_zones = {'Minimal': (0, 5), 'Mild': (5, 10), 'Moderate': (10, 15), 'Moderately Severe': (15, 20), 'Severe': (20, 27)}
phq9_colors = ['green', 'yellow', 'orange', 'red', 'darkred']
gad7_zones = {'Minimal': (0, 5), 'Mild': (5, 11), 'Moderate': (11, 15), 'Severe': (15, 21)}
gad7_colors = ['green', 'yellow', 'orange', 'red']
cgi_zones = {'Normal, not at all ill': (0, 1.5), 'Borderline ill': (1.5, 2.5), 'Midly ill': (2.5, 3.5), 'Moderately ill': (3.5, 4.5), 'Markedly ill': (4.5, 5.5), 'Severly ill': (5.5, 6.5), 'Among the most ill patients': (6.5, 7.5)}
cgi_colors = ['lightgreen', 'green', 'yellow', 'orange', 'red', 'darkred', 'black']
saps_sans_global_zones = {'Low Symptom Burden': (0,30), 'Moderate Symptom Burden': (30,60), 'Severe Symptom Burden': (60,101)}
saps_sans_global_colors = ['green', 'yellow', 'red']
saps_sans_symptoms_zones = {'Not at all': (0,0.5), 'Questionable': (0.5,1.5), 'Mild': (1.5,2.5), 'Moderate': (2.5,3.5), 'Marked': (3.5,4.5), 'Severe': (4.5, 5.5)}
saps_sans_symptoms_colors = ['lightgreen', 'green', 'yellow', 'orange', 'red', 'darkred']

# LOAD DATA

# Define the path to the folder containing the CSV files
sleep_data_folder_path = r'C:\Users\dylan\OneDrive - McGill University\McPsyt Lab Douglas\actigraphy\sleep_data'
activity_data_folder_path = r'C:\Users\dylan\OneDrive - McGill University\McPsyt Lab Douglas\actigraphy\activity_data'
redcap_data_folder_path = r'C:\Users\dylan\OneDrive - McGill University\McPsyt Lab Douglas\actigraphy\redcap_data'

# List all CSV files in the folder
sleep_files = [file for file in os.listdir(sleep_data_folder_path) if file.endswith('.csv')]
activity_files = [file for file in os.listdir(activity_data_folder_path) if file.endswith('.csv')]
redcap_files = [file for file in os.listdir(redcap_data_folder_path) if file.endswith('.xlsx')]

sleep_data = {}
activity_data = {}
redcap_data = {}


for sleep_file in sleep_files:
    sleep_file_name = os.path.splitext(sleep_file)[0]
    sleep_data[sleep_file_name] = pd.read_csv(os.path.join(sleep_data_folder_path, sleep_file))

for activity_file in activity_files:
    activity_file_name = os.path.splitext(activity_file)[0]
    activity_data[activity_file_name] = pd.read_csv(os.path.join(activity_data_folder_path, activity_file))

for redcap_file in redcap_files:
    redcap_file_name = os.path.splitext(redcap_file)[0]
    redcap_data[redcap_file_name] = pd.read_excel(os.path.join(redcap_data_folder_path, redcap_file))


list_participant_ids = [
    '001','002', '003', '004', '005', '006', '007', '008', '009', '010', '011', '012', '013', '014', '015'
]


for participant_index in list_participant_ids:

    print(f'PROCESSING PARTICIPANT DD_{participant_index}')

    max_digit_sleep = -1
    max_digit_activity = -1
    sleep_df = None
    activity_df = None
    redcap_df = None

    for sleep_file_name, df in sleep_data.items():
        if sleep_file_name.startswith(participant_index):
            
            # Extract the last character of the file name (assuming it ends with a digit)
            last_digit = int(sleep_file_name[-1])
            
            # Check if this digit is the largest found so far
            if last_digit > max_digit_sleep:
                max_digit_sleep = last_digit
                sleep_df = df
                sleep_df_name = sleep_file_name



    for activity_file_name, df in activity_data.items():
        if activity_file_name.startswith(participant_index):
            
            # Extract the last character of the file name (assuming it ends with a digit)
            last_digit = int(activity_file_name[-1])
            
            # Check if this digit is the largest found so far
            if last_digit > max_digit_activity:
                max_digit_activity = last_digit
                activity_df = df
                activity_df_name = activity_file_name

    
    for redcap_file_name, df in redcap_data.items():
        if redcap_file_name.endswith(participant_index):
            redcap_df = df
            redcap_df_name = redcap_file_name

            # Rename some columns
            redcap_df = redcap_df.rename(columns={
                'Date WHODAS': 'date_whodas',
                'WHODAS': 'whodas',
                'DAST10': 'dast10',
                'Date DAST10': 'date_dast10',
                'AUDIT': 'audit',
                'Date AUDIT': 'date_audit'
            })




    # If the patient doesn't have actigraphy data, but has redcap data
    if sleep_df is None and redcap_df is not None:

        # Convert to datetime
        redcap_df['date_phq9'] = pd.to_datetime(redcap_df['date_phq9'])
        redcap_df['date_gad7'] = pd.to_datetime(redcap_df['date_gad7'])
        redcap_df['date_cgi-s'] = pd.to_datetime(redcap_df['date_cgi-s'])
        redcap_df['date_saps-sans'] = pd.to_datetime(redcap_df['date_saps-sans'])
        redcap_df['date_whodas'] = pd.to_datetime(redcap_df['date_whodas'])
        redcap_df['date_dast10'] = pd.to_datetime(redcap_df['date_dast10'])
        redcap_df['date_audit'] = pd.to_datetime(redcap_df['date_audit'])


        # Initialize lists to hold valid min and max dates
        min_dates = []
        max_dates = []


        for col in ['date_phq9', 'date_gad7', 'date_cgi-s', 'date_saps-sans', 'date_whodas', 'date_dast10', 'date_audit']:
            if not redcap_df[col].isnull().all():  # Check if the column is not entirely null
                min_dates.append(redcap_df[col].min())
                max_dates.append(redcap_df[col].max())


        # Determine overall min and max date
        min_date = (min(min_dates) - timedelta(days=10)) if min_dates else None
        max_date = (max(max_dates) + timedelta(days=10)) if max_dates else None





        # Create a date range for x-ticks from min_date - 1 day to max_date + 1 day
        daily_date_range = pd.date_range(start=min_date - pd.Timedelta(days=1), end=max_date + pd.Timedelta(days=1), freq='D')
        weekly_date_range = pd.date_range(start=min_date - pd.Timedelta(days=1), end=max_date + pd.Timedelta(days=1), freq='W')

        # Define x-limits
        x_limits = (min_date - pd.Timedelta(days=1), max_date + pd.Timedelta(days=1))

        # Filter data within x-limits
        filtered_phq9_df = redcap_df[(redcap_df['date_phq9'] >= x_limits[0]) & (redcap_df['date_phq9'] <= x_limits[1])]

        # PHQ-9 Plot
        plt.figure(figsize=(9, 6))
        for idx, (zone, limits) in enumerate(phq9_zones.items()):
            plt.axhspan(limits[0], limits[1], color=phq9_colors[idx], alpha=0.3, label=zone)

        # Plot filtered data
        plt.plot(filtered_phq9_df['date_phq9'], filtered_phq9_df['phq_9'], marker='o', label='PHQ-9 Scores')

        # Add annotations for PHQ-9 values within the filtered data
        for i, score in enumerate(filtered_phq9_df['phq_9']):
            if pd.notna(score):
                plt.annotate(f'{score:.0f}', 
                            (filtered_phq9_df['date_phq9'].iloc[i], score), 
                            textcoords="offset points", 
                            xytext=(0, 10), ha='center')

        plt.ylabel('PHQ-9 Score')
        plt.ylim(0, 27)
        plt.xlim(x_limits)
        plt.xticks(ticks=daily_date_range, labels=[d.strftime('%Y-%m-%d') if d in weekly_date_range else '' for d in daily_date_range],rotation=90)
        plt.yticks(np.arange(0, 28, 3))
        plt.legend(loc='best')
        plt.tight_layout()
        plt.savefig(f'../individual_monthly_reports/{participant_index}/phq9_report_{participant_index}.png', dpi=300)



         # GAD-7 Plot
        plt.figure(figsize=(9, 6))
        for idx, (zone, limits) in enumerate(gad7_zones.items()):
            plt.axhspan(limits[0], limits[1], color=gad7_colors[idx], alpha=0.3, label=zone)
        plt.plot(redcap_df['date_gad7'], redcap_df['gad_7'], marker='o', label='GAD-7 Scores')

        # Add annotations for GAD-7 values
        for i, score in enumerate(redcap_df['gad_7']):
            if pd.notna(score):
                plt.annotate(f'{score:.0f}', (redcap_df['date_gad7'].iloc[i], score), textcoords="offset points", xytext=(0,10), ha='center')

        plt.ylabel('GAD-7 Score')
        plt.ylim(0, 21)
        plt.xlim(x_limits)
        plt.yticks(np.arange(0, 19, 3))
        plt.xticks(ticks=daily_date_range, labels=[d.strftime('%Y-%m-%d') if d in weekly_date_range else '' for d in daily_date_range], rotation=90)
        plt.legend(loc = 'best')
        plt.tight_layout()
        plt.savefig(f'../individual_monthly_reports/{participant_index}/gad7_report_{participant_index}.png', dpi=300)





        # CGI-S Plot
        plt.figure(figsize=(9, 6))
        for idx, (zone, limits) in enumerate(cgi_zones.items()):
            plt.axhspan(limits[0], limits[1], color=cgi_colors[idx], alpha=0.3, label=zone)
        plt.plot(redcap_df['date_cgi-s'], redcap_df['cgi_s'], marker='o', label='CGI-S Scores')

        # Add annotations for CGI values
        for i, score in enumerate(redcap_df['cgi_s']):
            if pd.notna(score):
                plt.annotate(f'{score:.0f}', (redcap_df['date_cgi-s'].iloc[i], score), textcoords="offset points", xytext=(0,10), ha='center')

        plt.ylabel('CGI-S Score')
        plt.ylim(0, 7)
        plt.xlim(x_limits)
        plt.yticks(np.arange(0, 8, 1))
        plt.xticks(ticks=daily_date_range, labels=[d.strftime('%Y-%m-%d') if d in weekly_date_range else '' for d in daily_date_range], rotation=90)
        plt.legend(loc = 'best')
        plt.tight_layout()
        plt.savefig(f'../individual_monthly_reports/{participant_index}/cgis_report_{participant_index}.png', dpi=300)





        ############ TABLE
        # Convert saps-sans date so it only keeps year, month and date (not time)
        redcap_df['date_saps-sans'] = pd.to_datetime(redcap_df['date_saps-sans']).dt.strftime('%Y-%m-%d')
        redcap_df['date_dast10'] = pd.to_datetime(redcap_df['date_dast10']).dt.strftime('%Y-%m-%d')
        redcap_df['date_audit'] = pd.to_datetime(redcap_df['date_audit']).dt.strftime('%Y-%m-%d')
        redcap_df['date_whodas'] = pd.to_datetime(redcap_df['date_whodas']).dt.strftime('%Y-%m-%d')

        saps_sans_symptoms = {
            'Date': redcap_df['date_saps-sans'],
            'Hallucinations': redcap_df['hallucinations'],
            'Delusion': redcap_df['delusions'],
            'Bizarre Behavior': redcap_df['bizarre_behavior'],
            'Apathy': redcap_df['apathy'],
            'Asociality': redcap_df['asociality'],
            'Attention': redcap_df['attention']
        }
        saps_sans_symptoms_df = pd.DataFrame(saps_sans_symptoms)

        saps_sans_global = {
            'Date': redcap_df['date_saps-sans'],
            'Total SAPS Score': redcap_df['total_saps'],
            'Total SANS Score': redcap_df['total_sans']
        }
        saps_sans_global_df = pd.DataFrame(saps_sans_global)

        dast_audit = {
            'DAST10 Score': redcap_df['dast10'],
            'Date DAST10': redcap_df['date_dast10'],
            'AUDIT': redcap_df['audit'],
            'Date AUDIT': redcap_df['date_audit']
        }
        dast_audit_df = pd.DataFrame(dast_audit)

        whodas = {
            'WHODAS': redcap_df['whodas'],
            'Date WHODAS': redcap_df['date_whodas']
        }
        whodas_df = pd.DataFrame(whodas)

        # File path
        file_path = f'../individual_monthly_reports/{participant_index}/table_report_{participant_index}.xlsx'


        if os.path.exists(file_path):
            os.remove(file_path)


        wb = Workbook()
        # wb.create_sheet(title='Sleep Data')
        # wb.create_sheet(title='Physical Activity')
        wb.create_sheet(title='SAPS-SANS Symptoms')
        wb.create_sheet(title='SAPS-SANS Global')
        wb.create_sheet(title='Drugs and Alcohol')
        wb.create_sheet(title='WHODAS')
        # Remove the default 'Sheet' created by openpyxl
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']
        wb.save(file_path)

        # Save sheets to the existing Excel file
        with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            saps_sans_symptoms_df.to_excel(writer, sheet_name='SAPS-SANS Symptoms', index=False)
            saps_sans_global_df.to_excel(writer, sheet_name='SAPS-SANS Global', index=False)
            dast_audit_df.to_excel(writer, sheet_name='Drugs and Alcohol', index=False)
            whodas_df.to_excel(writer, sheet_name='WHODAS', index=False)



        # Load the existing Excel workbook
        wb = load_workbook(file_path)
        symptoms_ws = wb['SAPS-SANS Symptoms']

        # Define color fills for severity levels
        fills = {
            0: PatternFill(start_color='CCFFCC', end_color='CCFFCC', fill_type='solid'),  # light green
            1: PatternFill(start_color='99FF99', end_color='99FF99', fill_type='solid'),  # green
            2: PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid'),  # yellow
            3: PatternFill(start_color='FFCC99', end_color='FFCC99', fill_type='solid'),  # orange
            4: PatternFill(start_color='FF6666', end_color='FF6666', fill_type='solid'),  # red
            5: PatternFill(start_color='990000', end_color='990000', fill_type='solid')   # dark red
        }

        # Apply color coding to severity levels in the SAPS-SANS Symptoms sheet
        for row in symptoms_ws.iter_rows(min_row=2, min_col=2, max_col=symptoms_ws.max_column):
            for cell in row:
                if cell.value in fills:  # Check if the cell value is a valid severity score
                    cell.fill = fills[cell.value]


        # Apply color coding to the SAPS-SANS Global sheet
        global_ws = wb['SAPS-SANS Global']

        # Define color fills for total score ranges (Global)
        global_fills = {
            'green': PatternFill(start_color='99FF99', end_color='99FF99', fill_type='solid'),  # 0-30
            'yellow': PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid'),  # 30-60
            'red': PatternFill(start_color='FF6666', end_color='FF6666', fill_type='solid')     # 60-100
        }

        # Color-code total scores for SAPS and SANS in SAPS-SANS Global sheet
        for row in global_ws.iter_rows(min_row=2, min_col=2, max_col=3):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    if 0 <= cell.value <= 30:
                        cell.fill = global_fills['green']
                    elif 30 < cell.value <= 60:
                        cell.fill = global_fills['yellow']
                    elif 60 < cell.value <= 100:
                        cell.fill = global_fills['red']


        # Access the 'SAPS-SANS Symptoms' sheet
        saps_sans_symptoms_ws.insert_rows(1)
        saps_sans_symptoms_ws.merge_cells('A1:G1')
        saps_sans_symptoms_ws['A1'] = 'SAPS-SANS Symptom Scores'
        saps_sans_symptoms_ws['A1'].font = Font(bold=True)
        saps_sans_symptoms_ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

        saps_sans_global_ws.insert_rows(1)
        saps_sans_global_ws.merge_cells('A1:C1')
        saps_sans_global_ws['A1'] = 'SAPS-SANS Global Scores'
        saps_sans_global_ws['A1'].font = Font(bold=True)
        saps_sans_global_ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

        dast_audit_wb = wb['Drugs and Alcohol']
        dast_audit_wb.insert_rows(1)
        dast_audit_wb.merge_cells('A1:D1')
        dast_audit_wb['A1'] = 'DAST10 and AUDIT Scores'
        dast_audit_wb['A1'].font = Font(bold=True)
        dast_audit_wb['A1'].alignment = Alignment(horizontal='center', vertical='center')

        whodas_df = wb['WHODAS']
        whodas_df.insert_rows(1)
        whodas_df['A1'] = 'WHODAS Score'
        whodas_df['A1'].font = Font(bold=True)
        whodas_df['A1'].alignment = Alignment(horizontal='center', vertical='center')


        



        # Save the changes to the same file
        wb.save(file_path)



















    if activity_df is not None or sleep_df is not None:

        print('THE FOLLOWING DATAFRAMES WILL BE USED TO CONSTRUCT THE GRAPHS:')
        print(sleep_df_name, activity_df_name, redcap_df_name)

        
        # Store mean onset time and mean rise time
        mean_onset = sleep_df.loc[sleep_df['Night.Starting'] == 'Mean', 'Sleep.Onset.Time'].iloc[0]
        mean_rise = sleep_df.loc[sleep_df['Night.Starting'] == 'Mean', 'Rise.Time'].iloc[0]
        print(mean_onset)
        print(mean_rise)

        # Remove Mean row now that we won't use it
        sleep_df = sleep_df[sleep_df['Night.Starting'] != 'Mean']

        if redcap_df is not None:
            # Convert to datetime
            redcap_df['date_phq9'] = pd.to_datetime(redcap_df['date_phq9'])
            redcap_df['date_gad7'] = pd.to_datetime(redcap_df['date_gad7'])
            redcap_df['date_cgi-s'] = pd.to_datetime(redcap_df['date_cgi-s'])
            redcap_df['date_saps-sans'] = pd.to_datetime(redcap_df['date_saps-sans'])
            redcap_df['date_whodas'] = pd.to_datetime(redcap_df['date_whodas'])
            redcap_df['date_dast10'] = pd.to_datetime(redcap_df['date_dast10'])
            redcap_df['date_audit'] = pd.to_datetime(redcap_df['date_audit'])
        sleep_df['Night.Starting'] = pd.to_datetime(sleep_df['Night.Starting'])
        

        # Create a Date column for activity_df based on the dates in the corresponding sleep_df
        activity_df['Date'] = pd.concat([pd.Series(sleep_df['Night.Starting'].iloc[0] - pd.Timedelta(days=1)), sleep_df['Night.Starting']]).reset_index(drop=True)
        activity_df['Date'] = pd.to_datetime(activity_df['Date'])

        # Remove earliest and latest date of activity data, as it is incomplete
        earliest_date = activity_df['Date'].min()
        latest_date = activity_df['Date'].max()
        print(f'The date {earliest_date} has been removed for participant {participant_index}')
        print(f'The date {latest_date} has been removed for participant {participant_index}')
        activity_df = activity_df[activity_df['Date'] != earliest_date]
        activity_df = activity_df[activity_df['Date'] != latest_date]
    
        # Convert Sleep Onset and Rise Time to datetime
        sleep_df['Sleep.Onset.Time'] = pd.to_datetime(sleep_df['Sleep.Onset.Time'], format='%H:%M', errors='coerce')
        sleep_df['Rise.Time'] = pd.to_datetime(sleep_df['Rise.Time'], format='%H:%M', errors='coerce')

        # Set a specific date (12-10-2003) for both times
        arbitrary_date = datetime.strptime('12-10-2003', '%d-%m-%Y')
        sleep_df['Sleep.Onset.Time'] = sleep_df['Sleep.Onset.Time'].apply(lambda x: x.replace(year=arbitrary_date.year, month=arbitrary_date.month, day=arbitrary_date.day) if pd.notnull(x) else x)
        sleep_df['Rise.Time'] = sleep_df['Rise.Time'].apply(lambda x: x.replace(year=arbitrary_date.year, month=arbitrary_date.month, day=arbitrary_date.day) if pd.notnull(x) else x)

        def adjust_sleep_time(row, date_comparator):
            time = row.time()
            if time < pd.Timestamp('06:00').time() and date_comparator.date() == pd.Timestamp('2003-10-12').date():
                return (date_comparator + pd.Timedelta(days=1)).replace(hour=time.hour, minute=time.minute)
            
            return date_comparator.replace(hour=time.hour, minute=time.minute)

        date_comparator = pd.Timestamp('2003-10-12 06:00')

        # Apply the adjustment to Sleep Time Onset and Rise Time columns
        sleep_df['Sleep.Onset.Time'] = sleep_df['Sleep.Onset.Time'].apply(lambda x: adjust_sleep_time(x, date_comparator))
        sleep_df['Rise.Time'] = sleep_df['Rise.Time'].apply(lambda x: adjust_sleep_time(x, date_comparator))

        onset_variability = sleep_df['Sleep.Onset.Time'].diff().abs().dt.total_seconds().mean() / 3600  # Convert to hours
        rise_variability = sleep_df['Rise.Time'].diff().abs().dt.total_seconds().mean() / 3600  # Convert to hours
        print(onset_variability)
        print(rise_variability)

        # Remove irrelevant rows
        # i.e. rows that have 0 on most columns and the "Mean" row
        activity_df = activity_df[(activity_df == 0).sum(axis=1) < 5]
        sleep_df = sleep_df[(sleep_df == 0).sum(axis=1) < 5]



        # Initialize lists to hold valid min and max dates
        min_dates = []
        max_dates = []

        # # ACTIVATE THIS SNIPPET IF YOU WANT TO INCLUDE REDCAP DATES INTO MIN AND MAX FOR X-AXIS
        # # Check each date column and add to lists if not empty
        # for col in ['date_phq9', 'date_gad7', 'date_cgi-s', 'date_saps-sans']:
        #     if not redcap_df[col].isnull().all():  # Check if the column is not entirely null
        #         min_dates.append(redcap_df[col].min())
        #         max_dates.append(redcap_df[col].max())

        if not sleep_df['Night.Starting'].isnull().all():  # Check if the column is not entirely null
            min_dates.append(sleep_df['Night.Starting'].min())
            max_dates.append(sleep_df['Night.Starting'].max())

        if not activity_df['Date'].isnull().all():  # Check if the column is not entirely null
            min_dates.append(activity_df['Date'].min())
            max_dates.append(activity_df['Date'].max())

        # Determine overall min and max date
        min_date = (min(min_dates) - timedelta(days=10)) if min_dates else None
        max_date = (max(max_dates) + timedelta(days=10)) if max_dates else None


        ##### THIS SNIPPET ALLOWS TO CONTROL X-AXIS INDIVIDUALLY
        if participant_index == '008':
            min_date = datetime.strptime("15/09/2024", "%d/%m/%Y")
            max_date = datetime.strptime("31/10/2024", "%d/%m/%Y")


        print(f'Min date: {min_date}')
        print(f'Max date: {max_date}')

        # Create a date range for x-ticks from min_date - 1 day to max_date + 1 day
        daily_date_range = pd.date_range(start=min_date - pd.Timedelta(days=1), end=max_date + pd.Timedelta(days=1), freq='D')
        weekly_date_range = pd.date_range(start=min_date - pd.Timedelta(days=1), end=max_date + pd.Timedelta(days=1), freq='W')

        # Define x-limits
        x_limits = (min_date - pd.Timedelta(days=1), max_date + pd.Timedelta(days=1))

       
        if redcap_df is not None:

            # PHQ-9 Plot
            # Filter data within x-limits
            filtered_phq9_df = redcap_df[(redcap_df['date_phq9'] >= x_limits[0]) & (redcap_df['date_phq9'] <= x_limits[1])]

            plt.figure(figsize=(9, 6))
            for idx, (zone, limits) in enumerate(phq9_zones.items()):
                plt.axhspan(limits[0], limits[1], color=phq9_colors[idx], alpha=0.3, label=zone)

            # Plot filtered data
            plt.plot(filtered_phq9_df['date_phq9'], filtered_phq9_df['phq_9'], marker='o', label='PHQ-9 Scores')

            # Add annotations for PHQ-9 values within the filtered data
            for i, score in enumerate(filtered_phq9_df['phq_9']):
                if pd.notna(score):
                    plt.annotate(f'{score:.0f}', 
                                (filtered_phq9_df['date_phq9'].iloc[i], score), 
                                textcoords="offset points", 
                                xytext=(0, 10), ha='center')

            plt.ylabel('PHQ-9 Score')
            plt.ylim(0, 27)
            plt.xlim(x_limits)
            plt.xticks(ticks=daily_date_range, labels=[d.strftime('%Y-%m-%d') if d in weekly_date_range else '' for d in daily_date_range],rotation=90)
            plt.yticks(np.arange(0, 28, 3))
            plt.legend(loc='best')
            plt.tight_layout()
            plt.savefig(f'../individual_monthly_reports/{participant_index}/phq9_report_{participant_index}.png', dpi=300)



            # GAD-7 Plot
            # Filter data within x-limits
            filtered_gad7_df = redcap_df[(redcap_df['date_gad7'] >= x_limits[0]) & (redcap_df['date_gad7'] <= x_limits[1])]
            plt.figure(figsize=(9, 6))
            for idx, (zone, limits) in enumerate(gad7_zones.items()):
                plt.axhspan(limits[0], limits[1], color=gad7_colors[idx], alpha=0.3, label=zone)
            plt.plot(filtered_gad7_df['date_gad7'], filtered_gad7_df['gad_7'], marker='o', color = '#CB6CE6', label='GAD-7 Scores')

            # Add annotations for GAD-7 values
            for i, score in enumerate(filtered_gad7_df['gad_7']):
                if pd.notna(score):
                    plt.annotate(f'{score:.0f}', (filtered_gad7_df['date_gad7'].iloc[i], score), textcoords="offset points", xytext=(0,10), ha='center')

            plt.ylabel('GAD-7 Score')
            plt.ylim(0, 21)
            plt.xlim(x_limits)
            plt.yticks(np.arange(0, 19, 3))
            plt.xticks(ticks=daily_date_range, labels=[d.strftime('%Y-%m-%d') if d in weekly_date_range else '' for d in daily_date_range], rotation=90)
            plt.legend(loc = 'best')
            plt.tight_layout()
            plt.savefig(f'../individual_monthly_reports/{participant_index}/gad7_report_{participant_index}.png', dpi=300)





            # CGI-S Plot
            # Filter data within x-limits
            filtered_cgis_df = redcap_df[(redcap_df['date_cgi-s'] >= x_limits[0]) & (redcap_df['date_cgi-s'] <= x_limits[1])]
            plt.figure(figsize=(9, 6))
            for idx, (zone, limits) in enumerate(cgi_zones.items()):
                plt.axhspan(limits[0], limits[1], color=cgi_colors[idx], alpha=0.3, label=zone)
            plt.plot(filtered_cgis_df['date_cgi-s'], filtered_cgis_df['cgi_s'], marker='o', color='#009E73', label='CGI-S Scores')

            # Add annotations for CGI values
            for i, score in enumerate(filtered_cgis_df['cgi_s']):
                if pd.notna(score):
                    plt.annotate(f'{score:.0f}', (filtered_cgis_df['date_cgi-s'].iloc[i], score), textcoords="offset points", xytext=(0,10), ha='center')

            plt.ylabel('CGI-S Score')
            plt.ylim(0, 7)
            plt.xlim(x_limits)
            plt.yticks(np.arange(0, 8, 1))
            plt.xticks(ticks=daily_date_range, labels=[d.strftime('%Y-%m-%d') if d in weekly_date_range else '' for d in daily_date_range], rotation=90)
            plt.legend(loc = 'best')
            plt.tight_layout()
            plt.savefig(f'../individual_monthly_reports/{participant_index}/cgis_report_{participant_index}.png', dpi=300)


            '''
            # SAPS and SANS Global Scores Plot
            plt.figure(figsize=(9, 6))
            saps_col = ['total_saps', 'total_sans']
            colors = ['red', 'blue']
            labels = ['SAPS Global Score', 'SANS Global Score']
            for col, color, label in zip(saps_col, colors, labels):
                plt.plot(redcap_df['date_saps-sans'], redcap_df[col], label=label, color=color, marker='o', markersize=5)

                #Annotating data points
                for i, score in enumerate(redcap_df[col]):
                    plt.annotate(f'{score:.0f}', (redcap_df['date_saps-sans'][i], redcap_df[col][i]),
                                textcoords="offset points", xytext=(0, 5), ha='center', fontsize=8)

            for idx, (zone, limits) in enumerate(saps_sans_global_zones.items()):
                plt.axhspan(limits[0], limits[1], color=saps_sans_global_colors[idx], alpha=0.3, label=zone)


            plt.ylabel('Global Score')
            plt.ylim(0, 101)
            plt.xlim(x_limits)
            plt.xticks(ticks=daily_date_range, labels=[d.strftime('%Y-%m-%d') if d in weekly_date_range else '' for d in daily_date_range], rotation=90)
            plt.yticks(np.arange(0, 101, 10))
            plt.legend(loc = 'best')
            plt.tight_layout()
            plt.savefig(f'../individual_monthly_reports/{participant_index}/total_saps_sans_report_{participant_index}.png', dpi=300)



            # SAPS and SANS Symptoms Plot
            plt.figure(figsize=(9, 6))
            saps_sans_sx = ['hallucinations', 'delusions', 'bizarre_behavior', 'apathy', 'asociality', 'attention']
            colors = ['red', 'orange', 'yellow', 'blue', 'purple', 'black']
            labels = ['Hallucinations', 'Delusions', 'Bizarre Behavior', 'Apathy', 'Asociality', 'Attention']
            annotation_markers = ['o', 's', '^', 'd', 'v', '*']

            jitter_amount = 0.2  # Adjust this value for more or less jitter

            for col, color, label, marker in zip(saps_sans_sx, colors, labels, annotation_markers):
                # Add jitter to the y-values
                jittered_y = redcap_df[col] + np.random.uniform(0, jitter_amount, size=redcap_df[col].shape)
                plt.plot(redcap_df['date_saps-sans'], jittered_y, label=label, color=color, marker=marker, markersize=5, markeredgecolor='black', markeredgewidth=0.2)



            for idx, (zone, limits) in enumerate(saps_sans_symptoms_zones.items()):
                plt.axhspan(limits[0], limits[1], color=saps_sans_symptoms_colors[idx], alpha=0.3, label=zone)

            plt.ylabel('Symptom Score')
            plt.ylim(0, 5.5)
            plt.xlim(x_limits)
            plt.xticks(ticks=daily_date_range, labels=[d.strftime('%Y-%m-%d') if d in weekly_date_range else '' for d in daily_date_range], rotation=90)
            plt.yticks(np.arange(0, 6, 1))
            plt.legend(loc = 'best')
            plt.tight_layout()
            plt.savefig(f'../individual_monthly_reports/{participant_index}/symptoms_saps_sans_report_{participant_index}.png', dpi=300)
            '''




        # SLEEP
        # Remove anomalies and convert sleep time to hours
        sleep_df = sleep_df[~((sleep_df['Total.Sleep.Time'] == 0) & (sleep_df['Total.Wake.Time'] == 0))]
        sleep_df['Total.Sleep.Time'] = sleep_df['Total.Sleep.Time'] / 3600
        sleep_df['Sleep.Difference'] = sleep_df['Total.Sleep.Time'].diff()
        sleep_variability = sleep_df['Sleep.Difference'].abs().mean()
        avg_sleep_efficiency = sleep_df['Sleep.Efficiency'].mean()


        # Total Sleep Time
        plt.figure(figsize=(9, 6))
        plt.plot(sleep_df['Night.Starting'], sleep_df['Total.Sleep.Time'], label='Sleep Duration', color='#5D3A9B', marker='o', markersize=5)
        avg_sleep = sleep_df['Total.Sleep.Time'].mean()
        plt.axhline(avg_sleep, linestyle='--', color='#5D3A9B', label=f'Avg Sleep Duration: {avg_sleep:.2f} hrs')
        # Add this in arguments of plt.legend() if you want the legend to be outside of the plot: loc='upper left', bbox_to_anchor=(1, 1)
        plt.legend(title=f'Daily Sleep Duration Variability: {sleep_variability:.2f} hrs')
        plt.ylabel('Time (Hours)')
        plt.xticks(ticks=daily_date_range, labels=[d.strftime('%Y-%m-%d') if d in weekly_date_range else '' for d in daily_date_range], rotation=90)
        plt.xlim(x_limits)
        plt.tight_layout()
        plt.savefig(f'../individual_monthly_reports/{participant_index}/sleep_time_report_{participant_index}.png', dpi=300)




        # # Sleep Efficiency
        # plt.figure(figsize=(9, 6))
        # plt.plot(sleep_df['Night.Starting'], sleep_df['Sleep.Efficiency'], label='Efficiency', color='red', marker='o', markersize=5)
        # avg_efficiency = sleep_df['Sleep.Efficiency'].mean()
        # plt.axhline(avg_efficiency, linestyle='--', color='red', label=f'Avg Efficiency: {avg_efficiency:.2f}%')
        # plt.ylabel('Efficiency (%)')
        # plt.yticks(np.arange(0, 101, 10))
        # plt.xticks(ticks=daily_date_range, labels=[d.strftime('%Y-%m-%d') if d in weekly_date_range else '' for d in daily_date_range], rotation=90)
        # plt.xlim(x_limits)
        # plt.legend()
        # plt.tight_layout()
        # plt.savefig(f'../individual_monthly_reports/{participant_index}/sleep_efficiency_report_{participant_index}.png', dpi=300)




        # # ACTIVITY
        # # Remove anomalies in activity data and convert to hours
        # activity_df = activity_df[~((activity_df['Sedentary'] == 0) & (activity_df['Light'] == 0) & (activity_df['Moderate'] == 0) & (activity_df['Vigorous'] == 0) & (activity_df['Steps'] == 0))]
        # activity_cols = ['Sedentary', 'Light', 'Moderate', 'Vigorous']
        # colors = ['blue', 'green', 'orange', 'red']  # Colors for each activity type
        # labels = ['Sedentary Activity', 'Light Physical Activity', 'Moderate Physical Activity', 'Vigorous Physical Activity']
        # activity_df[activity_cols] = activity_df[activity_cols] / 3600  # convert to hours

        # # Activity Plot for Sedentary, Light, Moderate, Vigorous
        # plt.figure(figsize=(9, 6))
        # for col, color in zip(activity_cols, colors):
        #     plt.plot(activity_df['Date'], activity_df[col], label=col, color=color, marker='o', markersize=5)
        #     avg_value = activity_df[col].mean()
        #     plt.axhline(avg_value, linestyle='--', color=color, label=f'Avg {col}: {avg_value:.2f} hrs')

        # # Add legend and labels
        # plt.legend(loc='upper left', bbox_to_anchor=(1, 1))
        # plt.ylabel('Time (Hours)')
        # plt.xticks(ticks=daily_date_range, labels=[d.strftime('%Y-%m-%d') if d in weekly_date_range else '' for d in daily_date_range], rotation=90)
        # plt.xlim(x_limits)
        # plt.tight_layout()
        # plt.savefig(f'../individual_monthly_reports/{participant_index}/activity_report_{participant_index}.png', dpi=300)




        # Steps Plot
        activity_df['Steps.Difference'] = activity_df['Steps'].diff()
        steps_variability = activity_df['Steps.Difference'].abs().mean()
        average_sedentary_activity = activity_df['Sedentary'].mean() / 3600
        average_light_activity = activity_df['Light'].mean() / 3600
        average_moderate_activity = activity_df['Moderate'].mean() / 3600
        average_vigorous_activity = activity_df['Vigorous'].mean() / 3600

        plt.figure(figsize=(9, 6))
        plt.plot(activity_df['Date'], activity_df['Steps'], label='Steps', color='#B61826', marker='o',  markersize=5)
        avg_steps = activity_df['Steps'].mean()
        plt.axhline(avg_steps, linestyle='--', color='#B61826', label=f'Avg Steps: {avg_steps:.0f}')
        plt.legend(title=f'Steps Variability: {steps_variability:.0f} steps')
        plt.ylabel('Steps')
        plt.xticks(ticks=daily_date_range, labels=[d.strftime('%Y-%m-%d') if d in weekly_date_range else '' for d in daily_date_range], rotation=90)
        plt.xlim(x_limits)
        plt.tight_layout()
        plt.savefig(f'../individual_monthly_reports/{participant_index}/steps_report_{participant_index}.png', dpi=300)


        # Table

        metrics_df = {
            'mean_sleep_time_onset': mean_onset,
            'sleep_time_onset_variability': f'{onset_variability:.2f}',
            'mean_rise_time': mean_rise,
            'rise_time_variability': f'{rise_variability:.2f}',
            'avg_sleep_efficiency': f'{sleep_df["Sleep.Efficiency"].mean():.2f}',
            'avg_sedentary_activity': f'{(activity_df["Sedentary"].mean() / 3600):.2f}',
            'avg_light_activity': f'{(activity_df["Light"].mean() / 3600):.2f}',
            'avg_moderate_activity': f'{(activity_df["Moderate"].mean() / 3600):.2f}',
            'avg_vigorous_activity': f'{(activity_df["Vigorous"].mean() / 3600):.2f}'
        }

        # Prepare Sleep Data sheet
        sleep_metrics = {
            'Category': ['Mean Sleep Time Onset', 'Sleep Time Onset Variability', 'Mean Rise Time', 'Rise Time Variability', 'Average Sleep Efficiency'],
            'Results': [metrics_df['mean_sleep_time_onset'], metrics_df['sleep_time_onset_variability'], metrics_df['mean_rise_time'], metrics_df['rise_time_variability'], metrics_df['avg_sleep_efficiency']]
        }
        sleep_metrics_df = pd.DataFrame(sleep_metrics)

        # Prepare Physical Activity sheet
        activity_metrics = {
            'Category': ['Average Sedentary Time', 'Average Light Activity', 'Average Moderate Activity', 'Average Vigorous Activity'],
            'Results': [metrics_df['avg_sedentary_activity'], metrics_df['avg_light_activity'], 
                        metrics_df['avg_moderate_activity'], metrics_df['avg_vigorous_activity']]
        }
        activity_metrics_df = pd.DataFrame(activity_metrics)

        if redcap_df is not None:

            # Convert saps-sans date so it only keeps year, month and date (not time)
            redcap_df['date_saps-sans'] = pd.to_datetime(redcap_df['date_saps-sans']).dt.strftime('%Y-%m-%d')
            redcap_df['date_dast10'] = pd.to_datetime(redcap_df['date_dast10']).dt.strftime('%Y-%m-%d')
            redcap_df['date_audit'] = pd.to_datetime(redcap_df['date_audit']).dt.strftime('%Y-%m-%d')
            redcap_df['date_whodas'] = pd.to_datetime(redcap_df['date_whodas']).dt.strftime('%Y-%m-%d')

            saps_sans_symptoms = {
                'Date': redcap_df['date_saps-sans'],
                'Hallucinations': redcap_df['hallucinations'],
                'Delusion': redcap_df['delusions'],
                'Bizarre Behavior': redcap_df['bizarre_behavior'],
                'Apathy': redcap_df['apathy'],
                'Asociality': redcap_df['asociality'],
                'Attention': redcap_df['attention']
            }
            saps_sans_symptoms_df = pd.DataFrame(saps_sans_symptoms)

            saps_sans_global = {
                'Date': redcap_df['date_saps-sans'],
                'Total SAPS Score': redcap_df['total_saps'],
                'Total SANS Score': redcap_df['total_sans']
            }
            saps_sans_global_df = pd.DataFrame(saps_sans_global)

            dast_audit = {
            'DAST10 Score': redcap_df['dast10'],
            'Date DAST10': redcap_df['date_dast10'],
            'AUDIT': redcap_df['audit'],
            'Date AUDIT': redcap_df['date_audit']
        }
            dast_audit_df = pd.DataFrame(dast_audit)

            whodas = {
                'WHODAS': redcap_df['whodas'],
                'Date WHODAS': redcap_df['date_whodas']
            }
            whodas_df = pd.DataFrame(whodas)

        # File path
        file_path = f'../individual_monthly_reports/{participant_index}/table_report_{participant_index}.xlsx'

        # Check if the file exists; if not, create it with empty sheets
        if not os.path.exists(file_path):
            wb = Workbook()
            wb.create_sheet(title='Sleep Data')
            wb.create_sheet(title='Physical Activity')
            
            if redcap_df is not None:
                wb.create_sheet(title='SAPS-SANS Symptoms')
                wb.create_sheet(title='SAPS-SANS Global')
                wb.create_sheet(title='Drugs and Alcohol')
                wb.create_sheet(title='WHODAS')
            # Remove the default 'Sheet' created by openpyxl
            if 'Sheet' in wb.sheetnames:
                del wb['Sheet']
            wb.save(file_path)

        # Save sheets to the existing Excel file
        with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            sleep_metrics_df.to_excel(writer, sheet_name='Sleep Data', index=False)
            activity_metrics_df.to_excel(writer, sheet_name='Physical Activity', index=False)
            
            if redcap_df is not None:
                saps_sans_symptoms_df.to_excel(writer, sheet_name='SAPS-SANS Symptoms', index=False)
                saps_sans_global_df.to_excel(writer, sheet_name='SAPS-SANS Global', index=False)
                dast_audit_df.to_excel(writer, sheet_name='Drugs and Alcohol', index=False)
                whodas_df.to_excel(writer, sheet_name='WHODAS', index=False)

        # Load the existing Excel workbook
        wb = load_workbook(file_path)

        if redcap_df is not None:
            symptoms_ws = wb['SAPS-SANS Symptoms']

            # Define color fills for severity levels
            fills = {
                0: PatternFill(start_color='CCFFCC', end_color='CCFFCC', fill_type='solid'),  # light green
                1: PatternFill(start_color='99FF99', end_color='99FF99', fill_type='solid'),  # green
                2: PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid'),  # yellow
                3: PatternFill(start_color='FFCC99', end_color='FFCC99', fill_type='solid'),  # orange
                4: PatternFill(start_color='FF6666', end_color='FF6666', fill_type='solid'),  # red
                5: PatternFill(start_color='990000', end_color='990000', fill_type='solid')   # dark red
            }

            # Apply color coding to severity levels in the SAPS-SANS Symptoms sheet
            for row in symptoms_ws.iter_rows(min_row=2, min_col=2, max_col=symptoms_ws.max_column):
                for cell in row:
                    if cell.value in fills:  # Check if the cell value is a valid severity score
                        cell.fill = fills[cell.value]

            # Apply color coding to the SAPS-SANS Global sheet
            global_ws = wb['SAPS-SANS Global']

            # Define color fills for total score ranges (Global)
            global_fills = {
                'green': PatternFill(start_color='99FF99', end_color='99FF99', fill_type='solid'),  # 0-30
                'yellow': PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid'),  # 30-60
                'red': PatternFill(start_color='FF6666', end_color='FF6666', fill_type='solid')     # 60-100
            }

            # Color-code total scores for SAPS and SANS in SAPS-SANS Global sheet
            for row in global_ws.iter_rows(min_row=2, min_col=2, max_col=3):
                for cell in row:
                    if isinstance(cell.value, (int, float)):
                        if 0 <= cell.value <= 30:
                            cell.fill = global_fills['green']
                        elif 30 < cell.value <= 60:
                            cell.fill = global_fills['yellow']
                        elif 60 < cell.value <= 100:
                            cell.fill = global_fills['red']

        # SET TITLES

        sleep_ws = wb['Sleep Data']
        sleep_ws.insert_rows(1)
        sleep_ws.merge_cells('A1:B1')
        sleep_ws['A1'] = 'KEY SLEEP METRICS'
        from openpyxl.styles import Font, Alignment
        sleep_ws['A1'].font = Font(bold=True)
        sleep_ws['A1'].alignment = Alignment(horizontal='center', vertical='center')


        activity_ws = wb['Physical Activity']
        activity_ws.insert_rows(1)
        activity_ws.merge_cells('A1:B1')
        activity_ws['A1'] = 'KEY ACTIVITY METRICS'
        activity_ws['A1'].font = Font(bold=True)
        activity_ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

        if redcap_df is not None:

            # Access the 'SAPS-SANS Symptoms' sheet
            saps_sans_symptoms_ws = wb['SAPS-SANS Symptoms']
            saps_sans_symptoms_ws.insert_rows(1)
            saps_sans_symptoms_ws.merge_cells('A1:G1')
            saps_sans_symptoms_ws['A1'] = 'SAPS-SANS Symptom Scores'
            saps_sans_symptoms_ws['A1'].font = Font(bold=True)
            saps_sans_symptoms_ws['A1'].alignment = Alignment(horizontal='center', vertical='center')


            saps_sans_global_ws = wb['SAPS-SANS Global']
            saps_sans_global_ws.insert_rows(1)
            saps_sans_global_ws.merge_cells('A1:C1')
            saps_sans_global_ws['A1'] = 'SAPS-SANS Global Scores'
            saps_sans_global_ws['A1'].font = Font(bold=True)
            saps_sans_global_ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

            dast_audit_wb = wb['Drugs and Alcohol']
            dast_audit_wb.insert_rows(1)
            dast_audit_wb.merge_cells('A1:D1')
            dast_audit_wb['A1'] = 'DAST10 and AUDIT Scores'
            dast_audit_wb['A1'].font = Font(bold=True)
            dast_audit_wb['A1'].alignment = Alignment(horizontal='center', vertical='center')

            whodas_df = wb['WHODAS']
            whodas_df.insert_rows(1)
            whodas_df['A1'] = 'WHODAS Score'
            whodas_df['A1'].font = Font(bold=True)
            whodas_df['A1'].alignment = Alignment(horizontal='center', vertical='center')



        # Save the changes to the same file
        wb.save(file_path)










    ################################################
    # THE FOLLOWING SNIPPETS CAN BE USED TO PLOT SLEEP ONSET TIME, SLEEP RISE TIME, NUMBER OF ACTIVE PERIODS, MEDIAN LENGTH OF ACTIVE PERIODS
    ################################################
    '''

        # Sleep Onset and Rise Time Plot

        # Remove any whitespace and replace empty strings with NaT
        sleep_df['Sleep.Onset.Time'] = sleep_df['Sleep.Onset.Time'].str.strip().replace('', pd.NaT)
        sleep_df['Rise.Time'] = sleep_df['Rise.Time'].str.strip().replace('', pd.NaT)

        # Convert Sleep Onset and Rise Time to datetime, setting an arbitrary date
        sleep_df['Sleep.Onset.Time'] = pd.to_datetime(sleep_df['Sleep.Onset.Time'], format='%H:%M', errors='coerce')
        sleep_df['Rise.Time'] = pd.to_datetime(sleep_df['Rise.Time'], format='%H:%M', errors='coerce')

        # Set a specific date (12-10-2003) for both times
        arbitrary_date = datetime.strptime('12-10-2003', '%d-%m-%Y')
        sleep_df['Sleep.Onset.Time'] = sleep_df['Sleep.Onset.Time'].apply(lambda x: x.replace(year=arbitrary_date.year, month=arbitrary_date.month, day=arbitrary_date.day) if pd.notnull(x) else x)
        sleep_df['Rise.Time'] = sleep_df['Rise.Time'].apply(lambda x: x.replace(year=arbitrary_date.year, month=arbitrary_date.month, day=arbitrary_date.day) if pd.notnull(x) else x)

        def adjust_sleep_time(row, date_comparator):
            time = row.time()
            if time < pd.Timestamp('06:00').time() and date_comparator.date() == pd.Timestamp('2003-10-12').date():
                return (date_comparator + pd.Timedelta(days=1)).replace(hour=time.hour, minute=time.minute)
            
            return date_comparator.replace(hour=time.hour, minute=time.minute)

        date_comparator = pd.Timestamp('2003-10-12 06:00')

        # Apply the adjustment to Sleep Time Onset and Rise Time columns
        sleep_df['Sleep.Onset.Time'] = sleep_df['Sleep.Onset.Time'].apply(lambda x: adjust_sleep_time(x, date_comparator))
        sleep_df['Rise.Time'] = sleep_df['Rise.Time'].apply(lambda x: adjust_sleep_time(x, date_comparator))

        # Convert Sleep Onset Times to hours for y-axis plotting
        # sleep_df['Sleep.Onset.Time'] = sleep_df['Sleep.Onset.Time'].dt.hour + sleep_df['Sleep.Onset.Time'].dt.minute / 60
        # sleep_df['Rise.Time'] = sleep_df['Rise.Time'].dt.hour + sleep_df['Rise.Time'].dt.minute / 60

        # Convert mean times to datetime on the same arbitrary date
        mean_onset = datetime.strptime(mean_onset, '%H:%M').replace(year=date_comparator.year, month=date_comparator.month, day=date_comparator.day)
        mean_rise = datetime.strptime(mean_rise, '%H:%M').replace(year=date_comparator.year, month=date_comparator.month, day=date_comparator.day)


        # Create a plot
        plt.figure(figsize=(12, 6))
        plt.plot(sleep_df['Night.Starting'], sleep_df[['Sleep.Onset.Time', 'Rise.Time']], marker='o', label=['Sleep Onset', 'Rise Time'])
        plt.axhline(y=mean_onset, color='blue', linestyle='--', label=f'Mean Sleep Onset Time: {mean_onset.time()}\nDay-to-Day Variability: {onset_variability:.2f} hours')
        plt.axhline(y=mean_rise, color='orange', linestyle='--', label=f'Mean Sleep Rise Time: {mean_rise.time()}\nDay-to-Day Variability: {rise_variability:.2f} hours')


        # Making the damn y axis
        start_time = datetime(2003, 10, 12, 6)  # 06:00 AM on 12-10-2003
        end_time = datetime(2003, 10, 13, 5)    # 06:00 AM on 13-10-2003
        time_range = [start_time + timedelta(hours=i) for i in range(24)]  # 24 hours + starting hour
        plt.yticks(time_range)  # Set the y-ticks to the time range
        plt.ylim(pd.Timestamp('2003-10-12 05:00:00'), pd.Timestamp('2003-10-13 06:00:00'))
        plt.gca().set_yticklabels([t.strftime('%H:%M') for t in time_range])

        plt.title('Sleep Onset and Rise Time')
        plt.xlabel('Date')
        plt.ylabel('Time')
        plt.xticks(ticks=daily_date_range, labels=[d.strftime('%Y-%m-%d') if d in weekly_date_range else '' for d in daily_date_range], rotation=90)
        plt.xlim(x_limits)
        plt.legend()
        plt.tight_layout()
        plt.savefig('../individual_monthly_reports/sleep_onset_and_rise_time_report.png', dpi=300)



        # Active periods during sleep Plot

        mean_num_active = sleep_df['Num.Active.Periods'].mean()
        mean_median_length = sleep_df['Median.Activity.Length'].mean()

        fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(10, 8), sharex=True, gridspec_kw={'hspace': 0})

        ax1.plot(sleep_df['Night.Starting'], sleep_df['Num.Active.Periods'], color='blue', marker='o', label='Number of Active Periods')
        ax1.axhline(mean_num_active, linestyle='--', color='blue', label=f'Avg Number of Active Periods: {mean_num_active:.0f}')
        ax1.set_ylabel('Number of Active Periods')
        ax1.tick_params(axis='y')
        ax1.legend(loc='upper left')
        ax1.set_title('Active Periods Count and Median Duration per Night of Sleep')

        ax2.plot(sleep_df['Night.Starting'], sleep_df['Median.Activity.Length'], color='red', marker='o', label='Median Length of Active Periods')
        ax2.axhline(mean_median_length, linestyle='--', color='red', label=f'Avg Median of Active Period Duration: {mean_median_length:.0f} minutes')
        ax2.set_ylabel('Median Activity Length (minutes)')
        ax2.tick_params(axis='y')
        ax2.set_xlabel('Date')
        ax2.legend()

        plt.xticks(ticks=daily_date_range, labels=[d.strftime('%Y-%m-%d') if d in weekly_date_range else '' for d in daily_date_range], rotation=90)
        plt.xlim(x_limits)

        # Adjust layout and save the figure
        fig.tight_layout()
        plt.savefig('../individual_monthly_reports/active_periods_during_sleep_report.png', dpi=300)
    '''